import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from collections import Counter
from collections import defaultdict

class ExcelDataProcessor:
    def __init__(self, csv_file_path, header=0):
        """初始化数据处理器"""
        self.csv_file_path = csv_file_path
        self.df = pd.read_csv(csv_file_path, header=header)

        self.COLUMN_MAPPING = {
            'serial_number': 2,            #序列号列 - 第3列
            'test_result': 7,               #测试结果列 - 第8列
            'station_id': 6,                  #站点ID列 - 第7列
            'start_date': 8,                #开始日期列 - 第9列
            'end_date': 9,                  #结束日期列 - 第10列
            'retest_type': 11,             #重测类型列 - 第12列
            'config': 4,                      #配置列 - 第13列
            'version': 10,          #WiPAS版本列 - 第11列
            'upper_limit': 0,               #上限行 - 第3行
            'lower_limit': 1                #下限行 - 第4列
        }
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'))

    def validate_column_mapping(self):
        """验证列映射是否有效"""
        max_cols = len(self.df.columns)
        for col_name, col_idx in self.COLUMN_MAPPING.items():
            if col_idx >= max_cols:
                raise ValueError(f"列索引 {col_idx} 超出数据范围，最大列数为 {max_cols}，列名: {col_name}")

    def get_column_data(self, column_name, start_row=5):
        """根据列名获取数据，支持指定起始行"""
        if column_name not in self.COLUMN_MAPPING:
            raise ValueError(f"未知的列名: {column_name}")
        col_idx = self.COLUMN_MAPPING[column_name]
        return self.df.iloc[start_row:, col_idx]

    def get_single_cell(self, column_name, row_idx):
        """获取单个单元格数据"""
        if column_name not in self.COLUMN_MAPPING:
            raise ValueError(f"未知的列名: {column_name}")
        col_idx = self.COLUMN_MAPPING[column_name]
        return self.df.iloc[row_idx, col_idx]

    def create_excel_with_sheets(self, filename):
        """创建一个包含多个工作表的Excel工作簿"""
        # 创建工作簿对象
        wb = Workbook()
        # 创建第一个工作表：Omnia Combined Auto
        self.create_all_config(wb)
        # 创建第二个工作表：
        self.each_config_write(wb)
        # 保存工作簿
        try:
            wb.save(filename)
            print(f"Excel工作簿 '{filename}' 创建成功，包含以下工作表：")
            for sheet in wb.sheetnames:
                print(f"- {sheet}")
            return True
        except Exception as e:
            print(f"创建Excel工作簿时出错: {e}")
            return False

    def create_all_config(self, wb):
        """创建For E-Mail工作簿"""
        ws1 = wb.active
        ws1.title = "Omnia Combined Auto"
        # 在第一行留空（实现空出一行的效果）
        # 原来第一行的内容下移一行
        Title = "T11 Pre_Proto_1 Combined Auto Daily Test Report"
        cell_a1 = ws1.cell(row=1, column=1, value=Title)
        cell_a1.alignment = Alignment(horizontal='center', vertical='center')
        cell_a1.border = self.thin_border
        cell_a1.font = Font(size=22, bold=True)
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        # 获取时间信息写入第二行
        daily_time_data = self.get_single_cell('end_date',3)
        if pd.notna(daily_time_data):
            # 使用pandas统一处理日期格式
            date_obj = pd.to_datetime(daily_time_data)
            daily_time_data = [date_obj.strftime('%Y/%m/%d')]
        else:
            daily_time_data = ['N/A']
        # print(daily_time_data)

        daily_time = daily_time_data[0] + " " + "04:00:00" + " " + "AM"
        # print(daily_time)
        cell_a2 = ws1.cell(row=2, column=1, value=daily_time)
        cell_a2.alignment = Alignment(horizontal='left', vertical='center')
        cell_a2.border = self.thin_border
        ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

        # 获取表头数据
        title_headers = self.get_column_data('config',3)
        title_headers = "T11 Pre_Proto_1_" + title_headers
        title_headers = list(dict.fromkeys(title_headers))  # config按顺序排列

        # 获取标题行数据（原数据Serial Number列）
        title_line_data1 = self.get_column_data('version',3)
        title_line_data1 = list(set(title_line_data1))

        title_line_data2 = self.get_single_cell('start_date',3)
        if pd.notna(title_line_data2):
            # 使用pandas统一处理日期格式
            date_obj = pd.to_datetime(title_line_data2)
            title_line_data2 = [date_obj.strftime('%Y/%m/%d')]
        else:
            title_line_data2 = ['N/A']
        # print(title_line_data2)

        title_line_data3 = self.get_single_cell('end_date',3)
        if pd.notna(title_line_data3):
            # 使用pandas统一处理日期格式
            date_obj = pd.to_datetime(title_line_data3)
            title_line_data3 = [date_obj.strftime('%Y/%m/%d')]
        else:
            title_line_data3 = ['N/A']

        # 在第三行添加表头内容（原第二行）
        header_cell = ws1.cell(row=3, column=1, value="Matrix Yield Statics")
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = self.thin_border
        # merge_cells：合并A1到C1单元格（第一行第一列到第一行第三列）
        ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

        for col, header in enumerate(title_headers, 3):
            cell = ws1.cell(row=3, column=col, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border

        # 写入标题行数据
        # column_data = [
        #     title_line_data1 * len(title_headers),
        #     [''] * len(title_headers),
        #     [''] * len(title_headers),
        # ]

        # 获取实际需要填充的config数量
        actual_config_count = len(self.get_column_data('config', 3).unique())

        # 获取WiPAS版本信息
        wipas_versions = list(set(self.get_column_data('version', 3).dropna()))

        # 创建列数据 - 只对应实际config数量
        column_data = [
            wipas_versions[:actual_config_count] + [''] * max(0, actual_config_count - len(wipas_versions)),
            [''] * actual_config_count,
            [''] * actual_config_count,
        ]


        # 定义第一列的内容
        first_column_values = ["Version",
                               "Start Test Date",
                               "Last Test Date",
                               ]
        for row_idx, data in enumerate(column_data, 4):  # 从第4行开始写入数据
            # 第一列添加指定内容
            first_col_value = first_column_values[row_idx - 4] if row_idx - 4 < len(first_column_values) else ""
            first_cell = ws1.cell(row=row_idx, column=1, value=first_col_value)
            first_cell.alignment = Alignment(horizontal='center', vertical='center')
            first_cell.border = self.thin_border
            # 合并每行的前二列单元格（A列到B列）
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)

            # 从第3列开始写入数据（因为前二列已合并）
            for col, value in enumerate(data, 3):
                cell = ws1.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
        # 对3～7行区域填充
        fill_36 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        for row in ws1['A3':'E6']:
            for cell in row:
                cell.fill = fill_36

        cell_a7 = ws1.cell(row=7, column=1, value="Total Input Quantity")
        cell_a7.alignment = Alignment(horizontal='center', vertical='center')
        cell_a7.border = self.thin_border

        cell_a8 = ws1.cell(row=8, column=1, value="Total Pass Quantity")
        cell_a8.alignment = Alignment(horizontal='center', vertical='center')
        cell_a8.border = self.thin_border
        fill_8 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # 对区域填充
        for row in ws1['A8':'E8']:
            for cell in row:
                cell.fill = fill_8

        cell_a9 = ws1.cell(row=9, column=1, value="1st Pass Quantity")
        cell_a9.alignment = Alignment(horizontal='center', vertical='center')
        cell_a9.border = self.thin_border

        cell_a10 = ws1.cell(row=10, column=1, value="Retest OK Quantity")
        cell_a10.alignment = Alignment(horizontal='center', vertical='center')
        cell_a10.border = self.thin_border

        cell_a11 = ws1.cell(row=11, column=1, value="Fail Quantity")
        cell_a11.alignment = Alignment(horizontal='center', vertical='center')
        cell_a11.border = self.thin_border
        fill_11 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        # 对区域填充
        for row in ws1['A11':'E11']:
            for cell in row:
                cell.fill = fill_11

        cell_a12 = ws1.cell(row=12, column=1, value="Total Pass Rate")
        cell_a12.alignment = Alignment(horizontal='center', vertical='center')
        cell_a12.border = self.thin_border
        fill_12 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # 对区域填充
        for row in ws1['A12':'E12']:
            for cell in row:
                cell.fill = fill_12

        # 第13行：1st Pass Rate
        cell_a13 = ws1.cell(row=13, column=1, value="1st Pass Rate")
        cell_a13.alignment = Alignment(horizontal='center', vertical='center')
        cell_a13.border = self.thin_border
        fill_13 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # 14行：Retest OK Quantity
        cell_a14 = ws1.cell(row=14, column=1, value="Retest OK Rate")
        cell_a14.alignment = Alignment(horizontal='center', vertical='center')
        cell_a14.border = self.thin_border

        # 第15行：Total Fail Rate
        cell_a15 = ws1.cell(row=15, column=1, value="Total Fail Rate")
        cell_a15.alignment = Alignment(horizontal='center', vertical='center')
        cell_a15.border = self.thin_border
        fill_15 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        # 对区域填充
        for row in ws1['A15':'R15']:
            for cell in row:
                cell.fill = fill_15

        # 原第16行下移至第17行：Test Waive Quantity
        cell_a16 = ws1.cell(row=16, column=1, value="Test Waive Quantity")
        cell_a16.alignment = Alignment(horizontal='center', vertical='center')
        cell_a16.border = self.thin_border

        # 写入B7单元格(计算总PASS个数)
        count_PASS = (self.get_column_data('test_result',3) == 'Pass').sum()
        cell_b7 = ws1.cell(row=7, column=2, value=count_PASS)
        cell_b7.alignment = Alignment(horizontal='center', vertical='center')
        cell_b7.border = self.thin_border

        # 写入B12单元格(计算FAIL>=3的个数)（原第11行）
        A = self.get_column_data('serial_number',3)
        B = self.get_column_data('test_result',3)
        data_dict = {'Serial_Number': A, 'Result': B}
        df = pd.DataFrame(data_dict)
        counts = df.groupby(['Serial_Number', 'Result']).size().reset_index(name='次数')
        # 筛选出FAIL次数>=3的记录
        fail_counts = counts[(counts['Result'] == 'Fail') & (counts['次数'] >= 3)]
        fail_ge_3_count = len(fail_counts)
        # print(fail_ge_3_count)
        cell_b11 = ws1.cell(row=11, column=2, value=fail_ge_3_count)
        cell_b11.alignment = Alignment(horizontal='center', vertical='center')
        cell_b11.border = self.thin_border

        # 写入B8单元格（总PASS数量减去误操作测Pass的个数）-- Total Pass Quantity
        cell_b8 = count_PASS    #- fail_ge_3_count  # 总数减去三次Fial的数量
        cell_b8 = ws1.cell(row=8, column=2, value=cell_b8)
        cell_b8.alignment = Alignment(horizontal='center', vertical='center')
        cell_b8.border = self.thin_border

        # 统计既有FAIL又有PASS记录的序列号数量--Retest OK Quantity
        pivot_table = df.pivot_table(index='Serial_Number', columns='Result', aggfunc='size', fill_value=0)
        # 确保两列都存在
        fail_col = 'Fail' if 'Fail' in pivot_table.columns else None
        pass_col = 'Pass' if 'Pass' in pivot_table.columns else None
        if fail_col and pass_col:
            both_fail_pass_count = len(pivot_table[(pivot_table[fail_col] > 0) & (pivot_table[pass_col] > 0)])
        else:
            both_fail_pass_count = 0
        cell_b10 = ws1.cell(row=10, column=2, value=both_fail_pass_count)
        cell_b10.alignment = Alignment(horizontal='center', vertical='center')
        cell_b10.border = self.thin_border

        # 1st Pass Quantity
        cell_b9 = cell_b8.value - both_fail_pass_count
        cell_b9 = ws1.cell(row=9, column=2, value=cell_b9)
        cell_b9.alignment = Alignment(horizontal='center', vertical='center')
        cell_b9.border = self.thin_border

        # Total Pass Rate
        cell_b12 = ws1.cell(row=12, column=2, value=cell_b8.value / cell_b7.value)
        cell_b12.number_format = '0.00%'
        cell_b12.alignment = Alignment(horizontal='center', vertical='center')
        cell_b12.border = self.thin_border

        # 1st Pass Rate
        cell_b13 = ws1.cell(row=13, column=2, value=cell_b9.value / cell_b7.value)
        cell_b13.number_format = '0.00%'
        cell_b13.alignment = Alignment(horizontal='center', vertical='center')
        cell_b13.border = self.thin_border

        # Retest OK Rate
        cell_b14 = ws1.cell(row=14, column=2, value=cell_b10.value / cell_b7.value)
        cell_b14.number_format = '0.00%'
        cell_b14.alignment = Alignment(horizontal='center', vertical='center')
        cell_b14.border = self.thin_border

        # Total Fail Rate
        cell_b15 = ws1.cell(row=15, column=2, value=fail_ge_3_count / cell_b7.value)
        cell_b15.number_format = '0.00%'
        cell_b15.alignment = Alignment(horizontal='center', vertical='center')
        cell_b15.border = self.thin_border

        # 调整列宽
        for col in range(1, len(title_headers) + 5):
            ws1.column_dimensions[get_column_letter(col)].width = 25

    def each_config_write(self, wb):
        """为每个唯一的config创建单独的工作表，布局与_create_all_config一致"""
        # 获取所有唯一的config名称
        all_configs = self.get_column_data('config',5).unique().tolist()

        for config_name in all_configs:
            if pd.notna(config_name):
                # 为每个config创建一个工作表
                ws = wb.create_sheet(title=str(config_name)[:31] + " " + "Config")  # Excel工作表名称限制为31个字符
                # 设置工作表标题
                title = f"T11 {config_name} Daily Test Report"
                cell_a1 = ws.cell(row=1, column=1, value=title)
                cell_a1.alignment = Alignment(horizontal='center', vertical='center')
                cell_a1.border = self.thin_border
                cell_a1.font = Font(size=22, bold=True)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

                # 获取时间信息写入第二行
                daily_time_data = self.get_single_cell('end_date',5)
                if pd.notna(daily_time_data):
                    # 使用pandas统一处理日期格式
                    date_obj = pd.to_datetime(daily_time_data)
                    daily_time_data = [date_obj.strftime('%Y/%m/%d')]
                else:
                    daily_time_data = ['N/A']

                daily_time = daily_time_data[0] + " " + "04:00:00" + " " + "AM"
                cell_a2 = ws.cell(row=2, column=1, value=daily_time)
                cell_a2.alignment = Alignment(horizontal='left', vertical='center')
                cell_a2.border = self.thin_border
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

                # 使用config_process函数计算统计数据
                config_processor = ConfigProcess(self.csv_file_path)
                result = config_processor.config_process(config_name)

                # 获取每日详细数据-将每日的数据信息写入到csv中
                daily_results = result.get('daily_result', {})
                # 为每一天的数据创建列并写入数据
                for i, (test_date, daily_data) in enumerate(daily_results.items()):
                    col_index = 3 + i  # 从第3列开始写入每日数据
                    # 写入各项每日统计数据
                    ws.cell(row=7, column=col_index, value=daily_data['total_input_qty']).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=7, column=col_index).border = self.thin_border
                    ws.cell(row=8, column=col_index, value=daily_data['unique_pass_qty']).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=8, column=col_index).border = self.thin_border
                    ws.cell(row=9, column=col_index, value=daily_data['first_pass_qty']).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=9, column=col_index).border = self.thin_border
                    ws.cell(row=10, column=col_index, value=daily_data['retest_ok_qty']).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=10, column=col_index).border = self.thin_border
                    ws.cell(row=11, column=col_index, value=daily_data['fail_qty']).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=11, column=col_index).border = self.thin_border
                    # 写入比率数据
                    # 需要转换百分比字符串为数值
                    for row_idx, rate_key in enumerate(
                            ['total_pass_rate', 'first_pass_rate', 'retest_ok_rate', 'fail_rate'], 12):
                        rate_value = daily_data[rate_key]
                        if isinstance(rate_value, str) and rate_value.endswith('%'):
                            rate_value = float(rate_value.rstrip('%')) / 100
                        cell = ws.cell(row=row_idx, column=col_index, value=rate_value)
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.thin_border

                # 获取每日时间数据
                daily_starts = result['daily_start'] if isinstance(result['daily_start'], list) else [
                    result['daily_start']]
                daily_ends = result['daily_end'] if isinstance(result['daily_end'], list) else [result['daily_end']]

                # 计算最大列数
                max_col = 2 + len(daily_starts)  # A,B列固定，从C列开始放置动态数据

                # 在第三行添加表头内容
                header_cell = ws.cell(row=3, column=1, value="Matrix Yield Statics")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = self.thin_border
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

                # 为每天的数据创建列标题
                for i in range(len(daily_starts)):
                    col_index = 3 + i
                    config_header = f"T11_Pre_Proto_1_{config_name}"
                    cell = ws.cell(row=3, column=col_index, value=config_header)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border

                # 定义第一列的内容
                first_column_values = ["WiPAS Version",
                                       "Start Test Date",
                                       "Last Test Date"]

                # 获取当前config的数据以提取WiPAS Version信息
                config_data = self.df[self.df.iloc[:,self.COLUMN_MAPPING['config']]== config_name].iloc[5:]
                title_line_data1 = config_data.iloc[:, self.COLUMN_MAPPING['version']].tolist() #.tolist()保留与否
                title_line_data1 = list(set(title_line_data1)) if title_line_data1 else ['N/A']

                # 写入第4-6行的数据（包括WiPAS Version等信息）
                # 第4行：WiPAS Version
                first_cell = ws.cell(row=4, column=1, value=first_column_values[0])
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.border = self.thin_border
                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)

                # 第4行：WiPAS Version
                first_cell = ws.cell(row=4, column=1, value=first_column_values[0])
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.border = self.thin_border
                ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)

                # 写入WiPAS Version对应的值到各列
                for i in range(len(daily_starts)):
                    col_index = 3 + i
                    # 填写实际的WiPAS Version值
                    cell_value = title_line_data1[0] if title_line_data1 and title_line_data1[0] != 'N/A' else ''
                    cell = ws.cell(row=4, column=col_index, value=cell_value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border

                # 第5行：Start Test Date
                first_cell = ws.cell(row=5, column=1, value=first_column_values[1])
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.border = self.thin_border
                ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)

                # 写入每日开始时间到各列
                for i in range(len(daily_starts)):
                    col_index = 3 + i
                    start_cell = ws.cell(row=5, column=col_index, value=daily_starts[i])
                    start_cell.alignment = Alignment(horizontal='center', vertical='center')
                    start_cell.border = self.thin_border

                # 第6行：Last Test Date
                first_cell = ws.cell(row=6, column=1, value=first_column_values[2])
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.border = self.thin_border
                ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)

                # 写入每日结束时间到各列
                for i in range(len(daily_ends)):
                    col_index = 3 + i
                    end_cell = ws.cell(row=6, column=col_index, value=daily_ends[i])
                    end_cell.alignment = Alignment(horizontal='center', vertical='center')
                    end_cell.border = self.thin_border

                # 对3～6行区域填充（与_create_all_config一致）
                fill_36 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                for row in ws[f'A3':f'{get_column_letter(max_col)}6']:
                    for cell in row:
                        cell.fill = fill_36

                # 添加统计行（与_create_all_config保持一致）
                cell_a7 = ws.cell(row=7, column=1, value="Total Input Quantity")
                cell_a7.alignment = Alignment(horizontal='center', vertical='center')
                cell_a7.border = self.thin_border

                cell_a8 = ws.cell(row=8, column=1, value="Total Pass Quantity")
                cell_a8.alignment = Alignment(horizontal='center', vertical='center')
                cell_a8.border = self.thin_border
                fill_8 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                for row in ws[f'A8':f'{get_column_letter(max_col)}8']:
                    for cell in row:
                        cell.fill = fill_8

                cell_a9 = ws.cell(row=9, column=1, value="1st Pass Quantity")
                cell_a9.alignment = Alignment(horizontal='center', vertical='center')
                cell_a9.border = self.thin_border

                cell_a10 = ws.cell(row=10, column=1, value="Retest OK Quantity")
                cell_a10.alignment = Alignment(horizontal='center', vertical='center')
                cell_a10.border = self.thin_border

                cell_a11 = ws.cell(row=11, column=1, value="Fail Quantity")
                cell_a11.alignment = Alignment(horizontal='center', vertical='center')
                cell_a11.border = self.thin_border
                fill_11 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                for row in ws[f'A11':f'{get_column_letter(max_col)}11']:
                    for cell in row:
                        cell.fill = fill_11

                cell_a12 = ws.cell(row=12, column=1, value="Total Pass Rate")
                cell_a12.alignment = Alignment(horizontal='center', vertical='center')
                cell_a12.border = self.thin_border
                fill_12 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                for row in ws[f'A12':f'{get_column_letter(max_col)}12']:
                    for cell in row:
                        cell.fill = fill_12

                # 第13行：1st Pass Rate
                cell_a13 = ws.cell(row=13, column=1, value="1st Pass Rate")
                cell_a13.alignment = Alignment(horizontal='center', vertical='center')
                cell_a13.border = self.thin_border

                # 14行：Retest OK Rate
                cell_a14 = ws.cell(row=14, column=1, value="Retest OK Rate")
                cell_a14.alignment = Alignment(horizontal='center', vertical='center')
                cell_a14.border = self.thin_border

                # 第15行：Total Fail Rate
                cell_a15 = ws.cell(row=15, column=1, value="Total Fail Rate")
                cell_a15.alignment = Alignment(horizontal='center', vertical='center')
                cell_a15.border = self.thin_border
                fill_15 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                for row in ws[f'A15':f'{get_column_letter(max_col)}15']:
                    for cell in row:
                        cell.fill = fill_15

                # 第16行：Test Waive Quantity
                cell_a16 = ws.cell(row=16, column=1, value="Test Waive Quantity")
                cell_a16.alignment = Alignment(horizontal='center', vertical='center')
                cell_a16.border = self.thin_border

                # 填入统计数据到第二列（与_create_all_config保持一致）
                # Total Input Quantity
                cell_b7 = ws.cell(row=7, column=2, value=result['total_input_qty'])
                cell_b7.alignment = Alignment(horizontal='center', vertical='center')
                cell_b7.border = self.thin_border
                # Total Pass Quantity
                cell_b8 = ws.cell(row=8, column=2, value=result['unique_pass_qty'])
                cell_b8.alignment = Alignment(horizontal='center', vertical='center')
                cell_b8.border = self.thin_border
                # 1st Pass Quantity
                cell_b9 = ws.cell(row=9, column=2, value=result['first_pass_qty'])
                cell_b9.alignment = Alignment(horizontal='center', vertical='center')
                cell_b9.border = self.thin_border
                # Retest OK Quantity
                cell_b10 = ws.cell(row=10, column=2, value=result['retest_ok_qty'])
                cell_b10.alignment = Alignment(horizontal='center', vertical='center')
                cell_b10.border = self.thin_border
                # Fail Quantity
                cell_b11 = ws.cell(row=11, column=2, value=result['fail_qty'])
                cell_b11.alignment = Alignment(horizontal='center', vertical='center')
                cell_b11.border = self.thin_border
                # 处理百分比数据
                if result['total_input_qty'] > 0:
                    # 转换百分比字符串为数值
                    total_pass_rate = float(result['total_pass_rate'].rstrip('%')) / 100
                    first_pass_rate = float(result['first_pass_rate'].rstrip('%')) / 100
                    retest_ok_rate = float(result['retest_ok_rate'].rstrip('%')) / 100
                    fail_rate = float(result['fail_rate'].rstrip('%')) / 100
                    # Total Pass Rate
                    cell_b12 = ws.cell(row=12, column=2, value=total_pass_rate)
                    cell_b12.number_format = '0.00%'
                    cell_b12.alignment = Alignment(horizontal='center', vertical='center')
                    cell_b12.border = self.thin_border
                    # 1st Pass Rate
                    cell_b13 = ws.cell(row=13, column=2, value=first_pass_rate)
                    cell_b13.number_format = '0.00%'
                    cell_b13.alignment = Alignment(horizontal='center', vertical='center')
                    cell_b13.border = self.thin_border
                    # Retest OK Rate
                    cell_b14 = ws.cell(row=14, column=2, value=retest_ok_rate)
                    cell_b14.number_format = '0.00%'
                    cell_b14.alignment = Alignment(horizontal='center', vertical='center')
                    cell_b14.border = self.thin_border
                    # Total Fail Rate
                    cell_b15 = ws.cell(row=15, column=2, value=fail_rate)
                    cell_b15.number_format = '0.00%'
                    cell_b15.alignment = Alignment(horizontal='center', vertical='center')
                    cell_b15.border = self.thin_border
                else:
                    # 如果没有数据，填入0
                    for row in range(12, 16):
                        cell = ws.cell(row=row, column=2, value=0)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.thin_border
                        cell.number_format = '0.00%'

                # 调整列宽
                for col in range(1, max_col + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 25

class ConfigProcess(ExcelDataProcessor):
    def __init__(self, csv_file_path, header=0, total_input_qty=None, total_pass_qty=None, first_pass_qty=None,
                 retest_ok_qty=None, fail_qty=None):
        # 只提取需要的属性
        super().__init__(csv_file_path, header)
        self.total_input_qty = total_input_qty  # 总的pass数量（包括误操作）
        self.total_pass_qty = total_pass_qty  # 实际的pass数量（去掉误操作）
        self.first_pass_qty = first_pass_qty  # 只测一次就pass的数量
        self.retest_ok_qty = retest_ok_qty  # AAB测试的数量
        self.fail_qty = fail_qty  # 测三次及以上Fail的数量
        self.last_total_pass_count = 0  # 新增实例变量存储最新计算的total_pass_count供后面计算重测百分比使用

    def config_process(self, config_name):  # 提取config,并计算各行数值
        config_extract = self.get_column_data('config',3)
        test_result = self.get_column_data('test_result',3)
        serial_number = self.get_column_data('serial_number',3)
        date_start = self.get_column_data('start_date',3)
        date_end = self.get_column_data('end_date',3)

        # 创建DataFrame
        df_ct = pd.DataFrame({
            'Start_Date': date_start.values,
            'End_Date': date_end.values,
            'Serial_Number': serial_number.values,
            'Test_Result': test_result.values,
            'Config': config_extract.values,
        })
        # 筛选特定config并统计
        df_filtered = df_ct[df_ct['Config'] == config_name].copy()

        # 处理日期数据并计算时间范围
        if not df_filtered.empty:
            # 转换日期列为datetime类型
            df_filtered['Start_Date'] = pd.to_datetime(df_filtered['Start_Date'],format='mixed',errors='coerce')
            df_filtered['End_Date'] = pd.to_datetime(df_filtered['End_Date'], format='mixed',errors='coerce')

            # 提取日期部分用于分组
            df_filtered['Test_Data'] = df_filtered['Start_Date'].dt.date
            # 分别计算每个config测试的每天最早开始时间和最晚结束时间
            daily_earliest_start = df_filtered.groupby('Test_Data')['Start_Date'].min()
            daily_latest_end = df_filtered.groupby('Test_Data')['End_Date'].max()
            # 格式化日期
            # 格式化每日日期为字符串列表
            daily_start_list = daily_earliest_start.dt.strftime(
                '%Y-%m-%d %H:%M:%S').tolist() if not daily_earliest_start.empty else ['N/A']
            daily_end_list = daily_latest_end.dt.strftime(
                '%Y-%m-%d %H:%M:%S').tolist() if not daily_latest_end.empty else ['N/A']

            # 获取该config整体的最早开始时间和最晚结束时间
            min_start_date = df_filtered['Start_Date'].min()
            max_end_date = df_filtered['End_Date'].max()
            # 格式化日期
            start_date_str = min_start_date.strftime('%Y-%m-%d') if pd.notna(min_start_date) else 'N/A'
            end_date_str = max_end_date.strftime('%Y-%m-%d') if pd.notna(max_end_date) else 'N/A'
        else:
            start_date_str = 'N/A'
            end_date_str = 'N/A'
            daily_start_list = 'N/A'
            daily_end_list = 'N/A'

        config_stats = df_filtered.groupby(['Test_Result']).size().reset_index(name='Count')
        # 提取总的pass和fail的数量（包括误差的）
        pass_count = int(config_stats[config_stats['Test_Result'] == 'Pass']['Count'].sum())
        fail_count = int(config_stats[config_stats['Test_Result'] == 'Fail']['Count'].sum())
        # 一、计算实际的pass数量（去掉误差的）
        total_pass_count = df_filtered[df_filtered['Test_Result'] == 'Pass']
        total_pass_count = total_pass_count['Serial_Number'].nunique()
        self.last_total_pass_count = total_pass_count  # 保存到实际变量供后面计算重测百分比

        # 二、计算只测一次就pass的数量
        first_test_records = df_filtered.groupby('Serial_Number').first().reset_index()
        one_pass_count = len(first_test_records[first_test_records['Test_Result'] == 'Pass'])
        # 三、计算AB/AAB测试pass的数量
        # 获取每个序列号的测试记录，按索引排序
        retest_groups = df_filtered.groupby('Serial_Number')['Test_Result'].apply(list)
        # 统计重测后通过的数量（至少有一次FAIL，最后的结果为PASS）
        # aab_pass_count = sum(1 for results in retest_groups if len(results)>1 and results[-1] == 'PASS' and 'FAIL' in results[:-1])
        aab_pass_count = 0
        for results in retest_groups:
            '''
            条件1：至少有2次测试
            条件2：最后一次测试结果是PASS
            条件3：前面的测试中有FAIL记录
            '''
            if len(results) > 1 and results[-1] == 'Pass' and 'Fail' in results[:-1]:
                aab_pass_count += 1
        # 四、计算三次都FAIL的数量
        fail_records = df_filtered[df_filtered['Test_Result'] == 'Fail']
        # 1、统计每个序列号的FAIL数量
        fail_counts = fail_records['Serial_Number'].value_counts()
        # 2、筛选出FAIl次数为3的序列号
        three_pass_sn = fail_counts[fail_counts == 3]
        # 3、统计数量
        all_pass_count = len(three_pass_sn)

        self.total_input_qty = pass_count
        self.total_pass_qty = total_pass_count
        self.first_pass_qty = one_pass_count
        self.retest_ok_qty = aab_pass_count
        self.fail_qty = all_pass_count
        self.total_pass_rate = self.total_pass_qty / self.total_input_qty if self.total_input_qty != 0 else 0
        self.first_pass_rate = self.first_pass_qty / self.total_input_qty if self.total_input_qty != 0 else 0
        self.retest_ok_rate = self.retest_ok_qty / self.total_input_qty if self.total_input_qty != 0 else 0
        self.fail_rate = self.fail_qty / self.total_input_qty if self.total_input_qty != 0 else 0

        # 按日期分组处理
        daily_results = {}
        for test_date, daily_data in df_filtered.groupby('Test_Data'):
            # 对每天的数据应用相同的计算逻辑
            daily_stats = daily_data.groupby(['Test_Result']).size().reset_index(name='Count')
            # 提取当天的pass和fail的数量
            pass_count = int(daily_stats[daily_stats['Test_Result'] == 'Pass']['Count'].sum())
            fail_count = int(daily_stats[daily_stats['Test_Result'] == 'Fail']['Count'].sum())
            # 计算当天的其他指标（复用原有逻辑）
            # 一、计算实际的pass数量（去掉误差的）
            total_pass_count = daily_data[daily_data['Test_Result'] == 'Pass']['Serial_Number'].nunique()
            # 二、计算只测一次就pass的数量
            first_test_records = daily_data.groupby('Serial_Number').first().reset_index()
            one_pass_count = len(first_test_records[first_test_records['Test_Result'] == 'Pass'])
            # 三、计算AB/AAB测试pass的数量
            retest_groups = daily_data.groupby('Serial_Number')['Test_Result'].apply(list)
            aab_pass_count = 0
            for results in retest_groups:
                '''
                条件1：至少有2次测试
                条件2：最后一次测试结果是PASS
                条件3：前面的测试中有FAIL记录
                '''
                if len(results) > 1 and results[-1] == 'Pass' and 'Fail' in results[:-1]:
                    aab_pass_count += 1

            # 四、计算三次都FAIL的数量
            fail_records = daily_data[daily_data['Test_Result'] == 'Fail']
            fail_counts = fail_records['Serial_Number'].value_counts()
            three_fail_sn = fail_counts[fail_counts == 3]
            all_fail_count = len(three_fail_sn)

            # 存储每日结果
            daily_results[test_date] = {
                'total_input_qty': pass_count,
                'unique_pass_qty': total_pass_count,
                'first_pass_qty': one_pass_count,
                'retest_ok_qty': aab_pass_count,
                'fail_qty': all_fail_count,
                # 计算每日的比率
                'total_pass_rate': f"{total_pass_count / pass_count:.2%}" if pass_count > 0 else "0.00%",
                'first_pass_rate': f"{one_pass_count / pass_count:.2%}" if pass_count > 0 else "0.00%",
                'retest_ok_rate': f"{aab_pass_count / pass_count:.2%}" if pass_count > 0 else "0.00%",
                'fail_rate': f"{all_fail_count / pass_count:.2%}" if pass_count > 0 else "0.00%"
            }

        return {
            # 'config': config_name,                                          #config名称
            # 'pass_count': pass_count,                                    #PASS数量
            # 'fail_count': fail_count,                                        #FAIL数量

            'daily_result': daily_results,  # 每天的各项计算

            'daily_start': daily_start_list,  # 每天config开始时间
            'daily_end': daily_end_list,  # 每天config结束时间

            'start_date': start_date_str,  # 整个config开始时间
            'end_date': end_date_str,  # 整个config结束时间

            'total_input_qty': self.total_input_qty,  # 总的pass数量（包括误操作）
            'unique_pass_qty': self.total_pass_qty,  # 实际的pass数量（去掉误操作）
            'first_pass_qty': self.first_pass_qty,  # 只测一次就pass的数量
            'retest_ok_qty': self.retest_ok_qty,  # ab/aab测试通过的pass的数量
            'fail_qty': self.fail_qty,  # 三次都Fail的pass数量
            # 返回格式化的百分比字符串
            'total_pass_rate': f"{self.total_pass_rate:.2%}" if self.total_input_qty else "0.00%",
            'first_pass_rate': f"{self.first_pass_rate:.2%}" if self.total_input_qty else "0.00%",
            'retest_ok_rate': f"{self.retest_ok_rate:.2%}" if self.total_input_qty else "0.00%",
            'fail_rate': f"{self.fail_rate:.2%}" if self.total_input_qty else "0.00%",

            'total_unique_pass_count': total_pass_count,  # 返回实际的pass数量，用于后面计算重测率
        }

    def config_write(self, excel_file, sheet_name, config_name, start_col=2, start_row=7):  # 将计算的数值依次写入表格不同列
        # 一、先将一列数据写入表格对应列；
        # 1、读取的数据CSV
        # 2、匹配（固定）到config列
        # 3、整列写入xlsx（填充）
        # 获取配置处理结果
        result = self.config_process(config_name)
        # 提取要写入的数据
        data_to_write = [
            result['total_input_qty'],
            result['unique_pass_qty'],
            result['first_pass_qty'],
            result['retest_ok_qty'],
            result['fail_qty'],
            result['total_pass_rate'],
            result['first_pass_rate'],
            result['retest_ok_rate'],
            result['fail_rate']
        ]
        # 使用openpyxl加载并修改Excel文件
        from openpyxl import load_workbook
        try:
            # 加载现有的Excel文件
            wb = load_workbook(excel_file)
            # 检查工作表是否存在
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print(f"工作表 '{sheet_name}' 不存在")
                return False

            # 写入日期信息
            start_date_cell = ws.cell(row=5, column=start_col, value=result['start_date'])
            end_date_cell = ws.cell(row=6, column=start_col, value=result['end_date'])
            # 设置日期单元格样式
            for cell in [start_date_cell, end_date_cell]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border

            # 从指定的起始行列位置开始写入数据（第3列的第8行到第17行）
            for i, value in enumerate(data_to_write):
                if start_row + i <= 17:  # 确保不超过第17行
                    cell = ws.cell(row=start_row + i, column=start_col, value=value)
                    # 设置单元格样式
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
            # 保存文件
            wb.save(excel_file)
            print(f"配置 '{config_name}' 的数据已成功写入 {excel_file} 的 '{sheet_name}' 工作表")
            return True
        except Exception as e:
            print(f"写入Excel文件时出错: {e}")
            return False


class RestestProcess(ExcelDataProcessor):
    def __init__(self, csv_file_path, header=0):
        # 只提取需要的属性
        super().__init__(csv_file_path, header)
        # 重测相关的属性
        self.retest_type = ''  # 重测的类型
        self.retest_count = 0  # 重测的数量
        self.retest_percentage = 0.0  # 重测的百分比
        self.station_id_count = []  # 重测的类型对应的ID和个数统计

    # 修改 RestestProcess 类中的 retest_process 方法
    def retest_process(self):
        '''
        1、提取retest[5:,11]列和Station ID [5:,6]列和SN[5:,2]列形成列表
        2、建立字典/列表    使用groupby函数
        3、统计重测类型、重测数量、重测比率、Station_ID、Station_Count
        4、建立表格并写入数据
        5、添加失败测试分析功能
        '''
        retest_type = self.get_column_data('retest_type',5)
        station_id = self.get_column_data('station_id',5)
        sn = self.get_column_data('serial_number',5)
        # 创建DataFrame
        df_ct = pd.DataFrame({
            'sn': sn,
            'retest_type': retest_type,
            'station_id': station_id
        })
        df_filtered = df_ct.dropna(how='any')  # 只保留三列都有值的行
        df_filtered = df_filtered.drop_duplicates(subset=['sn'], keep='last')
        df_filtered = df_filtered.reset_index(drop=True)  # 重制索引

        self.retest_type = set(df_filtered['retest_type'])  # 重测的类型
        self.retest_count = df_filtered['retest_type'].value_counts()  # 重测数量
        self.station_id_count = df_filtered.groupby(
            ['retest_type', 'station_id']).size()  # 按retest_type分组，统计每个类型下各station_id的计数

        # 创建 ConfigProcess 实例并获取 last_total_pass_count
        config_processor = ConfigProcess(self.csv_file_path)
        all_configs = self.get_column_data('config',5).unique().tolist()
        total_unique_pass_sum = 0
        # 遍历所有配置，累加每个配置的 last_total_pass_count
        for config_name in all_configs:
            if pd.notna(config_name):
                config_processor.config_process(config_name)  # 用 config_process 会更新 last_total_pass_count
                total_unique_pass_sum += config_processor.last_total_pass_count  # 使用 last_total_pass_count
        # 计算每个重测类型的百分比
        retest_percentage = {}
        for retest_type, count in self.retest_count.items():
            retest_percentage[retest_type] = count / total_unique_pass_sum if total_unique_pass_sum > 0 else 0

        # 添加失败测试分析功能 - 使用 demo1.py 的处理方式
        # 获取原始DataFrame的列名
        column_names = self.df.columns

        # 提取上下限数据
        upper_limit_row = self.df.iloc[self.COLUMN_MAPPING['upper_limit']]  # 上限行
        lower_limit_row = self.df.iloc[self.COLUMN_MAPPING['lower_limit']]  # 下限行

        # 构建一个新的DataFrame，包含上下限和所有数据列
        df_all = pd.concat([self.df.iloc[2:4], self.df.iloc[5:]]).copy()  # 合并上下限行和数据行
        df_all['original_index'] = df_all.index  # 保存原始索引

        # 提取最后的值，只对实际数据行进行处理（跳过上下限行）
        df_data_only = df_all.iloc[2:]  # 跳过前两行（上下限）
        df_data_only = df_data_only.dropna(how='any', subset=[self.df.columns[2], self.df.columns[11],
                                                              self.df.columns[12]])  # 只针对关键列检查空值，某一列存在空值，则删除该行
        df_data_only = df_data_only.drop_duplicates(subset=[self.df.columns[2]], keep='last')  # 基于SN列去重

        # 重新合并上下限行和处理后的数据行
        df_all = pd.concat([df_all.iloc[:2], df_data_only]).reset_index(drop=True)

        # 遍历df_all中的每一行，检查是否存在于retest_type中（跳过上下限行）
        result_values = []
        for index, row in df_all.iloc[2:].iterrows():  # 从第3行开始（索引2），跳过上下限行
            original_index = row['original_index']  # 获取原始索引
            retest_val = row.iloc[11]  # Retest Type列的值 (在df_all中是第12列，索引为11)

            if retest_val in column_names:
                # 如果Retest Type的值是列名之一，则获取该行该列的值
                value = self.df.at[original_index, retest_val]  # 从原始df获取交叉点的值
                result_values.append({
                    'row_index': original_index,
                    'column_name': retest_val,
                    'value': value
                })
            else:
                result_values.append(None)

        # 输出结果并进一步判断
        found_matches = [item for item in result_values if item is not None]

        # 初始化失败测试结果
        failing_tests_results = {}

        if found_matches:
            # 存储结果用于最后统一输出，按retest_type分组
            output_results = defaultdict(list)

            for match in found_matches:
                row_idx = match['row_index']
                col_name = match['column_name']
                value = match['value']
                # 获取对应的config值和retest_type值
                config_value = self.df.iloc[row_idx, 12]
                retest_type_value = self.df.iloc[row_idx, 11]

                # 判断是否为high或low
                result_category = None
                # 改进的数值类型检查
                try:
                    # 尝试将值转换为浮点数来判断是否为数值
                    if pd.notna(value):
                        # 先尝试转换为数值
                        numeric_value = float(value)
                        # 根据上下限判断是否为high/ low
                        try:
                            # 使用已定义的上下限变量
                            high_limit = upper_limit_row.get(col_name, None)
                            low_limit = lower_limit_row.get(col_name, None)
                            # 处理上下限数据
                            high_valid = False
                            low_valid = False
                            if pd.notna(high_limit):
                                try:
                                    high_limit = float(high_limit)
                                    high_valid = True
                                except (ValueError, TypeError):
                                    pass
                            if pd.notna(low_limit):
                                try:
                                    low_limit = float(low_limit)
                                    low_valid = True
                                except (ValueError, TypeError):
                                    pass
                            # 进行比较判断
                            if high_valid and numeric_value >= high_limit:
                                result_category = "high"
                            elif low_valid and numeric_value <= low_limit:
                                result_category = "low"
                            else:
                                if high_valid or low_valid:
                                    result_category = "normal"
                                else:
                                    result_category = "NA"
                        except Exception as e:
                            result_category = "NA"
                    else:
                        # 如果值为空，则标记为NA
                        result_category = "NA"
                except (ValueError, TypeError):
                    # 如果无法转换为数值，也标记为NA
                    result_category = "NA"
                # 按retest_type分组存储结果
                output_results[retest_type_value].append((config_value, result_category))

            # 保存失败测试结果
            failing_tests_results = output_results
        return {
            'retest_type': self.retest_type,  # 返回重测的类型
            'retest_count': self.retest_count,  # 返回重测的个数
            'station_id_count': self.station_id_count,  # 返回重测ID以及个数统计
            'retest_percentage': retest_percentage,  # 返回重测百分比
            'failing_tests_results': failing_tests_results  # 返回失败测试分析结果
        }

    def retest_write(self, excel_file, sheet_name, start_row=17, start_col=1):
        """
        将retest_process函数的返回值写入Excel文件
        Parameters:
        excel_file: Excel文件路径
        sheet_name: 工作表名称
        start_row=17 ：行开始
        start_col=1 ：列开始
        """
        # 用retest_process获取数据
        retest_data = self.retest_process()

        # 使用openpyxl加载并修改Excel文件
        from openpyxl import load_workbook

        try:
            # 加载现有的Excel文件
            wb = load_workbook(excel_file)
            # 检查工作表是否存在
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print(f"工作表 '{sheet_name}' 不存在")
                return False

            # 写入表头，增加"NO."和"Failing Tests"列
            header_row = start_row  # 从指定行开始写入表头
            headers = ["NO.", "Retest Type", "Retest Count", "Percentage", "Station ID", "Station Count",
                       "Failing Tests"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=header_row, column=start_col + i, value=header)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                cell.font = Font(bold=True)
            # 设置固定行高
            ws.row_dimensions[header_row].height = 15  # 设置行高为15

            # 写入数据 - 从表头下一行开始
            current_row = header_row + 1

            # 将多重索引数据转换为更容易处理的格式
            station_data = retest_data['station_id_count']

            # 按retest_type分组处理数据，按Retest Count降序排列
            retest_types = list(station_data.index.get_level_values(0).unique())
            # 按照retest_count进行排序
            retest_types.sort(key=lambda x: retest_data['retest_count'].get(x, 0), reverse=True)

            # 获取失败测试结果数据
            failing_tests_results = retest_data.get('failing_tests_results', {})

            # 分离特殊重测类型
            special_retest_type = "TxTestWithPowerSensor Test Error BT 1LE 2402 8 Ether_Scan"
            special_retest_types = [rt for rt in retest_types if rt == special_retest_type]
            normal_retest_types = [rt for rt in retest_types if rt != special_retest_type]

            # 序号计数器 - 在整个表格中连续递增
            row_number = 1

            # 先处理普通重测类型
            for retest_type in normal_retest_types:
                # 获取该retest_type下的所有station记录
                try:
                    subset = station_data.loc[retest_type]
                except KeyError:
                    # 如果找不到对应的retest_type，跳过
                    continue

                # 确保subset是Series格式
                if not isinstance(subset, pd.Series):
                    subset = pd.Series([subset], index=[subset.name] if hasattr(subset, 'name') else [None])

                # 获取该retest_type的总计数和百分比
                retest_count = retest_data['retest_count'].get(retest_type, 0)
                retest_percentage = retest_data['retest_percentage'].get(retest_type, 0)

                # 计算该retest_type有多少个station记录，用于合并单元格
                station_count = len(subset) if isinstance(subset, pd.Series) else 1

                # 处理每个station_id
                if isinstance(subset, pd.Series):
                    items = list(subset.items())
                else:
                    items = [(subset.name if hasattr(subset, 'name') else '', subset)]

                # 记录这一组数据的起始行
                group_start_row = current_row

                for idx, (station_id, station_count_val) in enumerate(items):
                    # 写入序号 - 使用全局递增的row_number，但只在第一行写入
                    if idx == 0:  # 只在每组的第一行写入序号
                        ws.cell(row=group_start_row, column=start_col, value=row_number).alignment = Alignment(
                            horizontal='center', vertical='center')
                        ws.cell(row=group_start_row, column=start_col).border = self.thin_border
                    else:
                        # 对于同一组的其他行，只添加边框而不写入序号值
                        ws.cell(row=current_row, column=start_col).border = self.thin_border

                    # 写入station_id和station_count（每行都写入）
                    modified_station_id = str(station_id)
                    if modified_station_id.startswith('LXKS_'):
                        modified_station_id = modified_station_id[5:]
                    ws.cell(row=current_row, column=start_col + 4, value=modified_station_id).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=start_col + 4).border = self.thin_border
                    ws.cell(row=current_row, column=start_col + 5, value=station_count_val).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=start_col + 5).border = self.thin_border

                    # 为所有单元格添加边框
                    for col_offset in range(7):  # 现在是7列（增加了NO.和Failing Tests列）
                        cell = ws.cell(row=current_row, column=start_col + col_offset)
                        if cell.border == Border():  # 只在还没有边框的情况下添加
                            cell.border = self.thin_border
                    current_row += 1

                # 在完成一组数据的写入后，合并需要合并的单元格
                if station_count > 0:
                    # 写入并合并NO.单元格（只在需要时合并）
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col,
                                       end_row=group_start_row + station_count - 1, end_column=start_col)

                    # 写入并合并retest_type单元格
                    retest_type_cell = ws.cell(row=group_start_row, column=start_col + 1, value=str(retest_type))
                    retest_type_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    retest_type_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 1,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 1)

                    # 写入并合并retest_count单元格
                    retest_count_cell = ws.cell(row=group_start_row, column=start_col + 2, value=retest_count)
                    retest_count_cell.alignment = Alignment(horizontal='center', vertical='center')
                    retest_count_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 2,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 2)

                    # 写入并合并percentage单元格
                    percentage_value = f"{retest_percentage:.2%}"  # 格式化为百分比
                    percentage_cell = ws.cell(row=group_start_row, column=start_col + 3, value=percentage_value)
                    percentage_cell.alignment = Alignment(horizontal='center', vertical='center')
                    percentage_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 3,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 3)

                    # 写入并合并Failing Tests单元格（在最后一列）
                    failing_tests_content = ""
                    if failing_tests_results and retest_type in failing_tests_results:
                        # 统计相同config和类型的组合数量
                        result_counter = Counter(failing_tests_results[retest_type])
                        # 按config名称排序输出
                        sorted_items = sorted(result_counter.items(), key=lambda x: x[0][0])
                        # 格式化为 config categoryxN 的形式
                        failing_tests_list = []
                        for (config_val, category), count in sorted_items:
                            failing_tests_list.append(f"{config_val} {category}x{count}")
                        failing_tests_content = "\n".join(failing_tests_list)

                    failing_tests_cell = ws.cell(row=group_start_row, column=start_col + 6, value=failing_tests_content)
                    failing_tests_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    failing_tests_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 6,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 6)

                # 每处理完一个retest_type组，序号递增1
                row_number += 1

            # 在普通数据和特殊数据之间添加一个空行
            current_row += 1

            # 为特殊重测类型添加表头（与原表头格式一致）
            special_header_row = current_row
            for i, header in enumerate(headers):
                cell = ws.cell(row=special_header_row, column=start_col + i, value=header)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                cell.font = Font(bold=True)
            # 设置固定行高
            ws.row_dimensions[special_header_row].height = 15  # 设置行高为15
            current_row += 1

            # 处理特殊重测类型
            for retest_type in special_retest_types:
                # 获取该retest_type下的所有station记录
                try:
                    subset = station_data.loc[retest_type]
                except KeyError:
                    # 如果找不到对应的retest_type，跳过
                    continue

                # 确保subset是Series格式
                if not isinstance(subset, pd.Series):
                    subset = pd.Series([subset], index=[subset.name] if hasattr(subset, 'name') else [None])

                # 获取该retest_type的总计数和百分比
                retest_count = retest_data['retest_count'].get(retest_type, 0)
                retest_percentage = retest_data['retest_percentage'].get(retest_type, 0)

                # 计算该retest_type有多少个station记录，用于合并单元格
                station_count = len(subset) if isinstance(subset, pd.Series) else 1

                # 处理每个station_id
                if isinstance(subset, pd.Series):
                    items = list(subset.items())
                else:
                    items = [(subset.name if hasattr(subset, 'name') else '', subset)]

                # 记录这一组数据的起始行
                group_start_row = current_row

                for idx, (station_id, station_count_val) in enumerate(items):
                    # 写入序号 - 使用全局递增的row_number，但只在第一行写入
                    if idx == 0:  # 只在每组的第一行写入序号
                        ws.cell(row=group_start_row, column=start_col, value=row_number).alignment = Alignment(
                            horizontal='center', vertical='center')
                        ws.cell(row=group_start_row, column=start_col).border = self.thin_border
                    else:
                        # 对于同一组的其他行，只添加边框而不写入序号值
                        ws.cell(row=current_row, column=start_col).border = self.thin_border

                    # 写入station_id和station_count（每行都写入）
                    modified_station_id = str(station_id)
                    if modified_station_id.startswith('LXKS_'):
                        modified_station_id = modified_station_id[5:]
                    ws.cell(row=current_row, column=start_col + 4, value=modified_station_id).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=start_col + 4).border = self.thin_border
                    ws.cell(row=current_row, column=start_col + 5, value=station_count_val).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=start_col + 5).border = self.thin_border

                    # 为所有单元格添加边框
                    for col_offset in range(7):  # 现在是7列（增加了NO.和Failing Tests列）
                        cell = ws.cell(row=current_row, column=start_col + col_offset)
                        if cell.border == Border():  # 只在还没有边框的情况下添加
                            cell.border = self.thin_border
                    current_row += 1

                # 在完成一组数据的写入后，合并需要合并的单元格
                if station_count > 0:
                    # 写入并合并NO.单元格（只在需要时合并）
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col,
                                       end_row=group_start_row + station_count - 1, end_column=start_col)

                    # 写入并合并retest_type单元格
                    retest_type_cell = ws.cell(row=group_start_row, column=start_col + 1, value=str(retest_type))
                    retest_type_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    retest_type_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 1,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 1)

                    # 写入并合并retest_count单元格
                    retest_count_cell = ws.cell(row=group_start_row, column=start_col + 2, value=retest_count)
                    retest_count_cell.alignment = Alignment(horizontal='center', vertical='center')
                    retest_count_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 2,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 2)

                    # 写入并合并percentage单元格
                    percentage_value = f"{retest_percentage:.2%}"  # 格式化为百分比
                    percentage_cell = ws.cell(row=group_start_row, column=start_col + 3, value=percentage_value)
                    percentage_cell.alignment = Alignment(horizontal='center', vertical='center')
                    percentage_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 3,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 3)

                    # 写入并合并Failing Tests单元格（在最后一列）
                    failing_tests_content = ""
                    if failing_tests_results and retest_type in failing_tests_results:
                        # 统计相同config和类型的组合数量
                        result_counter = Counter(failing_tests_results[retest_type])
                        # 按config名称排序输出
                        sorted_items = sorted(result_counter.items(), key=lambda x: x[0][0])
                        # 格式化为 config categoryxN 的形式
                        failing_tests_list = []
                        for (config_val, category), count in sorted_items:
                            failing_tests_list.append(f"{config_val} {category}x{count}")
                        failing_tests_content = "\n".join(failing_tests_list)

                    failing_tests_cell = ws.cell(row=group_start_row, column=start_col + 6, value=failing_tests_content)
                    failing_tests_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    failing_tests_cell.border = self.thin_border
                    if station_count > 1:
                        ws.merge_cells(start_row=group_start_row, start_column=start_col + 6,
                                       end_row=group_start_row + station_count - 1, end_column=start_col + 6)

                # 每处理完一个retest_type组，序号递增1
                row_number += 1

            # 在数据写入完成后设置数据行高
            for row_num in range(header_row + 1, current_row):
                ws.row_dimensions[row_num].height = 15

            # 保存文件
            wb.save(excel_file)
            print(f"重测数据已成功写入 {excel_file} 的 '{sheet_name}' 工作表")
            return True
        except Exception as e:
            print(f"写入Excel文件时出错: {e}")
            return False

    # 修改 each_retest_process 函数
    def each_retest_process(self, config_name):
        '''
        1、提取sn[5:,2]retest[5:,11]列和Station ID [5:,6]和config[5:,12]
        2、建立字典/列表    使用groupby函数
        3、统计每个config的重测类型、重测数量、重测比率、Station_ID_N
        4、建立表格并将数据写入每个config工作表
        5、添加失败测试分析功能
        '''
        sn_config = self.get_column_data('serial_number',5)
        retest_type_config = self.get_column_data('retest_type',5)
        station_id_config = self.get_column_data('station_id',5)
        retest_type_config_name = self.get_column_data('config',5)

        # 创建DataFrame
        df_ct_config = pd.DataFrame({
            'sn_config': sn_config,
            'retest_type_config': retest_type_config,
            'station_id_config': station_id_config,
            'retest_type_config_name': retest_type_config_name
        })

        df_filtered_config = df_ct_config.dropna()

        df_filtered_config = df_filtered_config.drop_duplicates(subset=['sn_config'], keep='last')

        df_filtered_config = df_filtered_config.reset_index(drop=True)  # 重制索引

        # 筛选特定config的数据
        df_filtered_config = df_filtered_config[df_filtered_config['retest_type_config_name'] == config_name]

        # 统计重测类型、数量和station ID计数
        retest_types = set(df_filtered_config['retest_type_config'])  # 重测的类型
        retest_counts = df_filtered_config['retest_type_config'].value_counts()  # 重测的个数
        # 按retest_type分组，统计每个类型下各station_id的计数
        station_id_counts = df_filtered_config.groupby(['retest_type_config', 'station_id_config']).size()

        # 创建 ConfigProcess 实例并获取 last_total_pass_count
        config_processor = ConfigProcess(self.csv_file_path)
        config_processor.config_process(config_name)
        total_unique_pass_sum = config_processor.last_total_pass_count

        # 计算每个重测类型的百分比
        retest_percentage = {}
        for retest_type, count in retest_counts.items():
            retest_percentage[retest_type] = count / total_unique_pass_sum if total_unique_pass_sum > 0 else 0

        # 添加失败测试分析功能
        # 获取原始DataFrame的列名
        column_names = self.df.columns
        # 提取上下限数据
        upper_limit_row = self.df.iloc[0]  # 上限行
        lower_limit_row = self.df.iloc[1]  # 下限行

        # 构建一个新的DataFrame，包含上下限和所有数据列
        df_all = pd.concat([self.df.iloc[0:2], self.df.iloc[3:]]).copy()  # 合并上下限行和数据行
        df_all['original_index'] = df_all.index  # 保存原始索引

        # 提取最后的值，只对实际数据行进行处理（跳过上下限行）
        df_data_only = df_all.iloc[2:]  # 跳过前两行（上下限）
        df_data_only = df_data_only.dropna(how='any', subset=[self.df.columns[2], self.df.columns[11],
                                                              self.df.columns[4]])  # 只针对关键列检查空值

        df_data_only = df_data_only.drop_duplicates(subset=[self.df.columns[2]], keep='last')  # 基于SN列去重
        # 重新合并上下限行和处理后的数据行
        df_all = pd.concat([df_all.iloc[:2], df_data_only]).reset_index(drop=True)

        # 遍历df_all中的每一行，检查是否存在于retest_type中（跳过上下限行）
        result_values = []
        for index, row in df_all.iloc[2:].iterrows():  # 从第3行开始（索引2），跳过上下限行
            original_index = row['original_index']  # 获取原始索引
            retest_val = row.iloc[11]  # Retest Type列的值 (在df_all中是第12列，索引为11)

            if retest_val in column_names:
                # 如果Retest Type的值是列名之一，则获取该行该列的值
                value = self.df.at[original_index, retest_val]  # 从原始df获取交叉点的值
                result_values.append({
                    'row_index': original_index,
                    'column_name': retest_val,
                    'value': value
                })
            else:
                result_values.append(None)

        # 输出结果并进一步判断
        found_matches = [item for item in result_values if item is not None]

        # 初始化失败测试结果
        failing_tests_results = {}

        if found_matches:
            # 存储结果用于最后统一输出，按retest_type分组
            output_results = defaultdict(list)

            for match in found_matches:
                row_idx = match['row_index']
                col_name = match['column_name']
                value = match['value']
                # 获取对应的config值和retest_type值 (确保只处理当前config的数据)
                config_value = self.df.iloc[row_idx, 12]
                retest_type_value = self.df.iloc[row_idx, 11]

                # 只处理当前config的数据
                if config_value == config_name:
                    # 判断是否为high或low
                    result_category = None
                    # 改进的数值类型检查
                    try:
                        # 尝试将值转换为浮点数来判断是否为数值
                        if pd.notna(value):
                            # 先尝试转换为数值
                            numeric_value = float(value)
                            # 根据上下限判断是否为high/ low
                            try:
                                # 使用已定义的上下限变量
                                high_limit = upper_limit_row.get(col_name, None)
                                low_limit = lower_limit_row.get(col_name, None)
                                # 处理上下限数据
                                high_valid = False
                                low_valid = False
                                if pd.notna(high_limit):
                                    try:
                                        high_limit = float(high_limit)
                                        high_valid = True
                                    except (ValueError, TypeError):
                                        pass
                                if pd.notna(low_limit):
                                    try:
                                        low_limit = float(low_limit)
                                        low_valid = True
                                    except (ValueError, TypeError):
                                        pass
                                # 进行比较判断
                                if high_valid and numeric_value >= high_limit:
                                    result_category = "high"
                                elif low_valid and numeric_value <= low_limit:
                                    result_category = "low"
                                else:
                                    if high_valid or low_valid:
                                        result_category = "normal"
                                    else:
                                        result_category = "NA"
                            except Exception as e:
                                result_category = "NA"
                        else:
                            # 如果值为空，则标记为NA
                            result_category = "NA"
                    except (ValueError, TypeError):
                        # 如果无法转换为数值，也标记为NA
                        result_category = "NA"
                    # 按retest_type分组存储结果
                    output_results[retest_type_value].append((config_value, result_category))

            # 保存失败测试结果
            failing_tests_results = output_results

        return {
            'retest_type_config': retest_types,
            'retest_count_config': retest_counts,
            'station_id_count_config': station_id_counts,
            'retest_percentage_config': retest_percentage,
            'failing_tests_results': failing_tests_results  # 新增失败测试分析结果
        }

    # 修改 each_retest_write 函数
    def each_retest_write(self, excel_file, sheet_name, start_row=17, start_col=1):
        """
        将each_retest_process函数的返回值写入Excel文件的特定config工作表中
        Parameters:
        excel_file: Excel文件路径
        sheet_name: 工作表名称
        start_row=17 ：行开始
        start_col=1 ：列开始
        """
        # 使用openpyxl加载并修改Excel文件
        from openpyxl import load_workbook

        try:
            # 加载现有的Excel文件
            wb = load_workbook(excel_file)
            # 检查工作表是否存在
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print(f"工作表 '{sheet_name}' 不存在")
                return False

            # 从sheet_name中提取config_name
            # 假设sheet_name格式为"{config_name} Config"
            if sheet_name.endswith(" Config"):
                config_name = sheet_name[:-7]  # 移除" Config"后缀
            else:
                config_name = sheet_name

            # 调用each_retest_process获取数据
            retest_data = self.each_retest_process(config_name)

            # 写入表头，增加"Failing Tests"列
            header_row = start_row  # 从指定行开始写入表头
            headers = ["NO.", "Retest Type", "Retest Count", "Percentage", "Station ID", "Station Count",
                       "Failing Tests"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=header_row, column=start_col + i, value=header)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                cell.font = Font(bold=True)
            # 设置固定行高
            ws.row_dimensions[header_row].height = 15  # 设置行高为15

            # 如果有数据则写入数据，否则只保留框架
            if retest_data['retest_type_config'] is not None and len(retest_data['retest_type_config']) > 0:
                # 写入数据 - 从表头下一行开始
                current_row = header_row + 1

                # 将多重索引数据转换为更容易处理的格式
                station_data = retest_data['station_id_count_config']

                # 按retest_type分组处理数据，按Retest Count降序排列
                retest_types = list(station_data.index.get_level_values(0).unique())
                # 按照retest_count进行排序
                retest_types.sort(key=lambda x: retest_data['retest_count_config'].get(x, 0), reverse=True)

                # 获取失败测试结果数据
                failing_tests_results = retest_data.get('failing_tests_results', {})


                # 分离特殊重测类型
                special_retest_type = "TxTestWithPowerSensor Test Error BT 1LE 2402 8 Ether_Scan"
                special_retest_types = [rt for rt in retest_types if rt == special_retest_type]
                normal_retest_types = [rt for rt in retest_types if rt != special_retest_type]

                # 序号计数器 - 在整个表格中连续递增
                row_number = 1

                # 先处理普通重测类型
                for retest_type in normal_retest_types:
                    # 获取该retest_type下的所有station记录
                    subset = station_data.loc[retest_type]

                    # 确保subset是Series格式
                    if not isinstance(subset, pd.Series):
                        subset = pd.Series([subset], index=[subset.name] if hasattr(subset, 'name') else [None])

                    # 获取该retest_type的总计数
                    retest_count = retest_data['retest_count_config'].get(retest_type, 0)
                    retest_percentage = retest_data['retest_percentage_config'].get(retest_type, 0)

                    # 计算该retest_type有多少个station记录，用于合并单元格
                    station_count = len(subset) if isinstance(subset, pd.Series) else 1

                    # 处理每个station_id
                    if isinstance(subset, pd.Series):
                        items = list(subset.items())
                    else:
                        items = [(subset.name if hasattr(subset, 'name') else '', subset)]

                    # 记录这一组数据的起始行
                    group_start_row = current_row

                    for idx, (station_id, station_count_val) in enumerate(items):
                        # 写入序号 - 使用全局递增的row_number，但只在第一行写入
                        if idx == 0:  # 只在每组的第一行写入序号
                            ws.cell(row=group_start_row, column=start_col, value=row_number).alignment = Alignment(
                                horizontal='center', vertical='center')
                            ws.cell(row=group_start_row, column=start_col).border = self.thin_border
                        else:
                            # 对于同一组的其他行，只添加边框而不写入序号值
                            ws.cell(row=current_row, column=start_col).border = self.thin_border

                        # 写入station_id和station_count（每行都写入）
                        modified_station_id = str(station_id)
                        if modified_station_id.startswith('LXKS_'):
                            modified_station_id = modified_station_id[5:]
                        ws.cell(row=current_row, column=start_col + 4, value=modified_station_id).alignment = Alignment(
                            horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=start_col + 4).border = self.thin_border
                        ws.cell(row=current_row, column=start_col + 5, value=station_count_val).alignment = Alignment(
                            horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=start_col + 5).border = self.thin_border

                        # 为所有单元格添加边框
                        for col_offset in range(7):  # 现在是7列（增加了NO.和Failing Tests列）
                            cell = ws.cell(row=current_row, column=start_col + col_offset)
                            if cell.border == Border():  # 只在还没有边框的情况下添加
                                cell.border = self.thin_border
                        current_row += 1

                    # 在完成一组数据的写入后，合并需要合并的单元格
                    if station_count > 0:
                        # 写入并合并NO.单元格（只在需要时合并）
                        if station_count > 1:
                            ws.merge_cells(start_row=group_start_row, start_column=start_col,
                                           end_row=group_start_row + station_count - 1, end_column=start_col)

                        # 写入并合并retest_type单元格
                        retest_type_cell = ws.cell(row=group_start_row, column=start_col + 1, value=str(retest_type))
                        retest_type_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        retest_type_cell.border = self.thin_border
                        if station_count > 1:
                            ws.merge_cells(start_row=group_start_row, start_column=start_col + 1,
                                           end_row=group_start_row + station_count - 1, end_column=start_col + 1)

                        # 写入并合并retest_count单元格
                        retest_count_cell = ws.cell(row=group_start_row, column=start_col + 2, value=retest_count)
                        retest_count_cell.alignment = Alignment(horizontal='center', vertical='center')
                        retest_count_cell.border = self.thin_border
                        if station_count > 1:
                            ws.merge_cells(start_row=group_start_row, start_column=start_col + 2,
                                           end_row=group_start_row + station_count - 1, end_column=start_col + 2)

                        # 写入并合并percentage单元格
                        percentage_value = f"{retest_percentage:.2%}"  # 格式化为百分比
                        percentage_cell = ws.cell(row=group_start_row, column=start_col + 3, value=percentage_value)
                        percentage_cell.alignment = Alignment(horizontal='center', vertical='center')
                        percentage_cell.border = self.thin_border
                        if station_count > 1:
                            ws.merge_cells(start_row=group_start_row, start_column=start_col + 3,
                                           end_row=group_start_row + station_count - 1, end_column=start_col + 3)

                        # 写入并合并Failing Tests单元格（在最后一列）
                        failing_tests_content = ""
                        if failing_tests_results and retest_type in failing_tests_results:
                            # 统计相同config和类型的组合数量
                            result_counter = Counter(failing_tests_results[retest_type])
                            # 按config名称排序输出
                            sorted_items = sorted(result_counter.items(), key=lambda x: x[0][0])
                            # 格式化为 config categoryxN 的形式
                            failing_tests_list = []
                            for (config_val, category), count in sorted_items:
                                failing_tests_list.append(f"{config_val} {category}x{count}")
                            failing_tests_content = "\n".join(failing_tests_list)

                        failing_tests_cell = ws.cell(row=group_start_row, column=start_col + 6,
                                                     value=failing_tests_content)
                        failing_tests_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        failing_tests_cell.border = self.thin_border
                        if station_count > 1:
                            ws.merge_cells(start_row=group_start_row, start_column=start_col + 6,
                                           end_row=group_start_row + station_count - 1, end_column=start_col + 6)

                    # 每处理完一个retest_type组，序号递增1
                    row_number += 1

                # 在普通数据和特殊数据之间添加一个空行
                current_row += 1

                # 为特殊重测类型添加表头（与原表头格式一致）
                if special_retest_types:  # 只有当存在特殊重测类型时才添加表头
                    special_header_row = current_row
                    # 写入表头，增加"Failing Tests"列
                    headers = ["NO.", "Retest Type", "Retest Count", "Percentage", "Station ID", "Station Count",
                               "Failing Tests"]
                    for i, header in enumerate(headers):
                        cell = ws.cell(row=special_header_row, column=start_col + i, value=header)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.thin_border
                        cell.font = Font(bold=True)
                    # 设置固定行高
                    ws.row_dimensions[special_header_row].height = 15  # 设置行高为15
                    current_row += 1

                    # 处理特殊重测类型
                    for retest_type in special_retest_types:
                        # 获取该retest_type下的所有station记录
                        subset = station_data.loc[retest_type]

                        # 确保subset是Series格式
                        if not isinstance(subset, pd.Series):
                            subset = pd.Series([subset], index=[subset.name] if hasattr(subset, 'name') else [None])

                        # 获取该retest_type的总计数
                        retest_count = retest_data['retest_count_config'].get(retest_type, 0)
                        retest_percentage = retest_data['retest_percentage_config'].get(retest_type, 0)

                        # 计算该retest_type有多少个station记录，用于合并单元格
                        station_count = len(subset) if isinstance(subset, pd.Series) else 1

                        # 处理每个station_id
                        if isinstance(subset, pd.Series):
                            items = list(subset.items())
                        else:
                            items = [(subset.name if hasattr(subset, 'name') else '', subset)]

                        # 记录这一组数据的起始行
                        group_start_row = current_row

                        for idx, (station_id, station_count_val) in enumerate(items):
                            # 写入序号 - 使用全局递增的row_number，但只在第一行写入
                            if idx == 0:  # 只在每组的第一行写入序号
                                ws.cell(row=group_start_row, column=start_col, value=row_number).alignment = Alignment(
                                    horizontal='center', vertical='center')
                                ws.cell(row=group_start_row, column=start_col).border = self.thin_border
                            else:
                                # 对于同一组的其他行，只添加边框而不写入序号值
                                ws.cell(row=current_row, column=start_col).border = self.thin_border

                            # 写入station_id和station_count（每行都写入）
                            modified_station_id = str(station_id)
                            if modified_station_id.startswith('LXKS_'):
                                modified_station_id = modified_station_id[5:]
                            ws.cell(row=current_row, column=start_col + 4,
                                    value=modified_station_id).alignment = Alignment(
                                horizontal='center', vertical='center')
                            ws.cell(row=current_row, column=start_col + 4).border = self.thin_border
                            ws.cell(row=current_row, column=start_col + 5,
                                    value=station_count_val).alignment = Alignment(
                                horizontal='center', vertical='center')
                            ws.cell(row=current_row, column=start_col + 5).border = self.thin_border

                            # 为所有单元格添加边框
                            for col_offset in range(7):  # 现在是7列（增加了NO.和Failing Tests列）
                                cell = ws.cell(row=current_row, column=start_col + col_offset)
                                if cell.border == Border():  # 只在还没有边框的情况下添加
                                    cell.border = self.thin_border
                            current_row += 1

                        # 在完成一组数据的写入后，合并需要合并的单元格
                        if station_count > 0:
                            # 写入并合并NO.单元格（只在需要时合并）
                            if station_count > 1:
                                ws.merge_cells(start_row=group_start_row, start_column=start_col,
                                               end_row=group_start_row + station_count - 1, end_column=start_col)

                            # 写入并合并retest_type单元格
                            retest_type_cell = ws.cell(row=group_start_row, column=start_col + 1,
                                                       value=str(retest_type))
                            retest_type_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            retest_type_cell.border = self.thin_border
                            if station_count > 1:
                                ws.merge_cells(start_row=group_start_row, start_column=start_col + 1,
                                               end_row=group_start_row + station_count - 1, end_column=start_col + 1)

                            # 写入并合并retest_count单元格
                            retest_count_cell = ws.cell(row=group_start_row, column=start_col + 2, value=retest_count)
                            retest_count_cell.alignment = Alignment(horizontal='center', vertical='center')
                            retest_count_cell.border = self.thin_border
                            if station_count > 1:
                                ws.merge_cells(start_row=group_start_row, start_column=start_col + 2,
                                               end_row=group_start_row + station_count - 1, end_column=start_col + 2)

                            # 写入并合并percentage单元格
                            percentage_value = f"{retest_percentage:.2%}"  # 格式化为百分比
                            percentage_cell = ws.cell(row=group_start_row, column=start_col + 3, value=percentage_value)
                            percentage_cell.alignment = Alignment(horizontal='center', vertical='center')
                            percentage_cell.border = self.thin_border
                            if station_count > 1:
                                ws.merge_cells(start_row=group_start_row, start_column=start_col + 3,
                                               end_row=group_start_row + station_count - 1, end_column=start_col + 3)

                            # 写入并合并Failing Tests单元格（在最后一列）
                            failing_tests_content = ""
                            if failing_tests_results and retest_type in failing_tests_results:
                                # 统计相同config和类型的组合数量
                                result_counter = Counter(failing_tests_results[retest_type])
                                # 按config名称排序输出
                                sorted_items = sorted(result_counter.items(), key=lambda x: x[0][0])
                                # 格式化为 config categoryxN 的形式
                                failing_tests_list = []
                                for (config_val, category), count in sorted_items:
                                    failing_tests_list.append(f"{config_val} {category}x{count}")
                                failing_tests_content = "\n".join(failing_tests_list)

                            failing_tests_cell = ws.cell(row=group_start_row, column=start_col + 6,
                                                         value=failing_tests_content)
                            failing_tests_cell.alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                            failing_tests_cell.border = self.thin_border
                            if station_count > 1:
                                ws.merge_cells(start_row=group_start_row, start_column=start_col + 6,
                                               end_row=group_start_row + station_count - 1, end_column=start_col + 6)

                        # 每处理完一个retest_type组，序号递增1
                        row_number += 1

                # 在数据写入完成后设置数据行高
                for row_num in range(header_row + 1, current_row):
                    ws.row_dimensions[row_num].height = 15

            # 只修改后两列的列宽设置
            ws.column_dimensions[get_column_letter(start_col + 4)].width = 20  # Station ID 列
            ws.column_dimensions[get_column_letter(start_col + 5)].width = 20  # Station Count 列
            ws.column_dimensions[get_column_letter(start_col + 6)].width = 25  # Failing Tests 列

            # 保存文件
            wb.save(excel_file)
            print(f"配置 '{config_name}' 的重测数据框架已成功写入 {excel_file} 的 '{sheet_name}' 工作表")
            return True
        except Exception as e:
            print(f"写入Excel文件时出错: {e}")
            return False

def main():
    """主函数"""
    # 创建数据处理器实例
    processor = ExcelDataProcessor('cwd_data.csv')
    # 生成包含多个工作表的Excel文件
    processor.create_excel_with_sheets('T11_CWB.xlsx')
    config_processor = ConfigProcess('cwd_data.csv')
    config_processor.create_excel_with_sheets('T11_CWB.xlsx')

    # 调用retest_process函数
    retest_processor = RestestProcess('cwd_data.csv')
    retest_processor.retest_write('T11_CWB.xlsx', 'Omnia Combined Auto')

    # 获取所有唯一的config名称
    all_configs = config_processor.get_column_data('config',5).unique().tolist()
    # 批量写入所有配置数据
    for i, config_name in enumerate(all_configs):
        if pd.notna(config_name):
            config_processor.config_write('T11_CWB.xlsx', 'Omnia Combined Auto',
                                          config_name, start_col=3 + i)
            # 调用每个配置的重测处理方法并写入数据
            sheet_name = str(config_name)[:31] + " " + "Config"  # 与each_config_write中创建的工作表名称保持一致
            retest_processor.each_retest_write('T11_CWB.xlsx',
                                               sheet_name)


if __name__ == "__main__":
    main()
